"""Quartalsweise IR-Berichte der Modekonzerne (Zalando, Inditex, H&M).

WICHTIG: Alle hier erhobenen KPIs sind KONZERNWEITE Werte aus offiziellen
IR-Veroeffentlichungen - KEINE Standort-/Filialdaten (die sind nicht oeffentlich
verfuegbar, siehe README).

Zwei Extraktionswege, je nachdem was ein Unternehmen anbietet:

1. XLSX-Weg (aktuell nur Zalando, config.ir_reports.companies.<key>.financials_xlsx_title
   gesetzt): Zalando veroeffentlicht pro Quartal eine strukturierte Excel-Datei
   ("Financials XLS") mit mehreren Sheets (Group key figures, Segment performance
   B2C/B2B, ...). Das ist deutlich zuverlaessiger als Text aus einem PDF zu regexen
   und liefert nebenbei die KOMPLETTE Quartalshistorie seit Q1/2020 in einem
   Rutsch (statt nur den aktuellen Wert). Siehe _fetch_xlsx_company().

2. PDF-Regex-Weg (Fallback, wenn kein financials_xlsx_title konfiguriert ist,
   z.B. Inditex/H&M): IR-Seite laden, neuesten Quartals-PDF-Link finden,
   Text extrahieren, konfigurierte Regex-KPIs suchen (nur der aktuelle Wert
   pro Lauf). Siehe _fetch_pdf_company().

In beiden Faellen wird mindestens der Berichtslink immer gespeichert
(data["ir_reports"]), auch wenn die KPI-Extraktion scheitert.
"""
import io
import re
from datetime import date

from common import add_point, http_get, now_iso, upsert_series

PDF_LINK_RE = re.compile(r'href="([^"]+\.pdf[^"]*)"', re.IGNORECASE)
QUARTER_HINT = re.compile(r"(q[1-4]|quarter|interim|halbjahr|nine[- ]month|trading update)", re.IGNORECASE)
QUARTER_COL_RE = re.compile(r"Q([1-4])\s*/\s*(\d{2})")
NUM_RE = re.compile(r"-?\d+(?:\.\d+)?")
_SUBROW_KEYWORDS = {"b2c", "b2b", "reconciliation"}

# corporate.zalando.com blockt den generischen common.USER_AGENT (nicht-Browser-UA)
# mit 403 Forbidden (WAF/Bot-Schutz) - mit einem browserartigen User-Agent klappt
# derselbe Request problemlos (lokal verifiziert). Analog zum muenchen-CKAN-503-Fix
# in radverkehr.py.
_BROWSER_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Accept-Language": "de-DE,de;q=0.9,en-US;q=0.8,en;q=0.7",
}


def _current_quarter_label():
    today = date.today()
    q = (today.month - 1) // 3 + 1
    q -= 1  # Berichte beziehen sich i.d.R. auf das Vorquartal
    year = today.year
    if q == 0:
        q, year = 4, year - 1
    return f"{year}-Q{q}"


def _quarter_to_isodate(year, quarter):
    end = {1: "03-31", 2: "06-30", 3: "09-30", 4: "12-31"}[quarter]
    return f"{year}-{end}"


def _find_pdf(ir_url):
    """Generischer Fallback: ersten PDF-Link mit Quartals-Hinweis auf der IR-Seite nehmen."""
    html = http_get(ir_url).text
    links = PDF_LINK_RE.findall(html)
    if not links:
        return None
    scored = sorted(links, key=lambda u: 0 if QUARTER_HINT.search(u) else 1)
    return _absolutize(scored[0], ir_url)


def _find_link_by_title(html, title_substr, base_url):
    """Link ueber sein title-Attribut finden (Zalandos Downloadbereich beschriftet
    jeden Link stabil, z.B. title="Q1 2026 Financials XLS" - der Dateiname aendert
    sich jedes Quartal, das title-Suffix nicht)."""
    m = re.search(
        r'title="([^"]*' + re.escape(title_substr) + r'[^"]*)"[^>]*href="([^"]+)"',
        html, re.IGNORECASE,
    )
    if not m:
        return None
    return _absolutize(m.group(2), base_url)


def _absolutize(url, base_url):
    if url.startswith("/"):
        from urllib.parse import urljoin
        return urljoin(base_url, url)
    return url


def _num(v):
    """Zellwert robust in float wandeln - manche Spalten liefern Strings wie
    '4.7x' (Excel-Zahlenformat mit Suffix) statt reiner Zahlen."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        m = NUM_RE.match(v.strip().replace(",", "."))
        if m:
            return float(m.group(0))
    return None


def _clean_label(raw):
    return re.sub(r"\*+$", "", str(raw).strip()).strip()


def _label_unit(label):
    """Grobe Heuristik: Einheit aus der letzten Klammer im Label extrahieren
    (z.B. 'Revenue (in m EUR)' -> 'Mio. EUR'). Margin-Zeilen bekommen '%'."""
    if "margin" in label.lower():
        return "%"
    m = re.search(r"\(([^)]*)\)\s*$", label)
    if not m:
        return ""
    inner = m.group(1)
    if not re.search(r"EUR|SEK|USD|\bm\b|%|x\b", inner, re.IGNORECASE):
        return ""
    unit = inner.replace("in m ", "Mio. ").replace("(m)", "Mio.")
    return unit


def _parse_xlsx_sheet(ws):
    """Ein Sheet der Zalando-'Financials XLS' -> [(label, unit, iso_date, value), ...].
    Findet die Kopfzeile (enthaelt 'Q1/20' o.ae.) automatisch, nimmt NUR echte
    Quartalsspalten (keine Halbjahr-/Jahres-Summenspalten - die haben 'H1'/'FY'
    statt 'Q1'..'Q4' im Spaltentext und werden vom QUARTER_COL_RE gar nicht erst
    getroffen)."""
    header_row = None
    quarter_cols = []  # [(col_idx, iso_date), ...]
    for r in range(1, min(15, ws.max_row) + 1):
        found = []
        for c in range(1, ws.max_column + 1):
            val = ws.cell(r, c).value
            if not val:
                continue
            m = QUARTER_COL_RE.search(str(val))
            if m:
                q, yy = int(m.group(1)), int(m.group(2))
                found.append((c, _quarter_to_isodate(2000 + yy, q)))
        if len(found) >= 4:  # eindeutig die Kopfzeile, nicht ein Zufallstreffer
            header_row = r
            quarter_cols = found
            break
    if header_row is None:
        return []

    out = []
    parent_label, parent_margin = None, False
    for r in range(header_row + 1, ws.max_row + 1):
        raw_label = ws.cell(r, 1).value
        if not raw_label or not isinstance(raw_label, str):
            continue
        cleaned = _clean_label(raw_label)
        if not cleaned or cleaned.lower().startswith(("*", "results of operations", "other key figures")):
            continue
        # Segment-Sheet: Zeilen "B2C"/"B2B"/"Reconciliation" sind Unterzeilen der
        # zuletzt gesehenen Hauptmetrik (z.B. "Revenue (in m EUR)") und muessen mit
        # ihr kombiniert werden, sonst landen GMV-B2C, Revenue-B2C, EBIT-B2C etc.
        # faelschlich alle in derselben Serie "B2C".
        if cleaned.lower() in _SUBROW_KEYWORDS and parent_label:
            label = f"{parent_label} – {cleaned}"
            unit = _label_unit(parent_label)
            is_margin = parent_margin
        else:
            label = cleaned
            unit = _label_unit(label)
            is_margin = "margin" in label.lower()
            parent_label, parent_margin = label, is_margin
        any_value = False
        for col, iso_date in quarter_cols:
            val = _num(ws.cell(r, col).value)
            if val is None:
                continue
            any_value = True
            out.append((label, unit, iso_date, val * 100 if is_margin else val))
        if not any_value:
            continue  # reine Text-/Trennzeile, keine KPI
    return out


def _fetch_xlsx_company(data, key, comp, errors):
    name = comp.get("name", key)
    reports = data.setdefault("ir_reports", {})
    html = http_get(comp["ir_url"], headers=_BROWSER_HEADERS).text
    xlsx_url = _find_link_by_title(html, comp["financials_xlsx_title"], comp["ir_url"])
    presentation_url = None
    if comp.get("presentation_title"):
        presentation_url = _find_link_by_title(html, comp["presentation_title"], comp["ir_url"])

    entry = reports.setdefault(key, {})
    entry.pop("kpis", None)  # Relikt aus dem alten PDF-Regex-Weg, hier nicht mehr genutzt
    entry.update({
        "name": name, "report_url": presentation_url or xlsx_url or comp["ir_url"],
        "ir_url": comp["ir_url"], "checked": now_iso(), "scope": "konzern",
    })
    if not xlsx_url:
        entry["quarter"] = _current_quarter_label()
        errors.append(("ir_reports", f"{name}: 'Financials XLS'-Link nicht gefunden (Seitenstruktur geaendert?)"))
        return

    import openpyxl
    xlsx_bytes = http_get(xlsx_url, headers=_BROWSER_HEADERS, timeout=90).content
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)

    n_kpis, n_points, latest_q, earliest_q = 0, 0, None, None
    for sheet_name in comp.get("xlsx_sheets", wb.sheetnames):
        if sheet_name not in wb.sheetnames:
            errors.append(("ir_reports", f"{name}: Sheet '{sheet_name}' nicht in {xlsx_url} gefunden"))
            continue
        rows = _parse_xlsx_sheet(wb[sheet_name])
        by_kpi = {}
        for label, unit, iso_date, val in rows:
            by_kpi.setdefault(label, []).append((unit, iso_date, val))
        for label, points in by_kpi.items():
            unit = points[0][0]
            sid = f"ir_{key}_" + re.sub(r"[^a-z0-9]+", "_", label.lower())[:50].strip("_")
            s = upsert_series(
                data, sid, label=f"{name}: {label}", frequency="quarterly",
                unit=unit, scope="konzern",
                source=f"{name} Investor Relations (Financials XLS, {sheet_name})",
                source_url=xlsx_url,
            )
            for _, iso_date, val in points:
                add_point(s, iso_date, val, "quarterly")
                latest_q = max(latest_q, iso_date) if latest_q else iso_date
                earliest_q = min(earliest_q, iso_date) if earliest_q else iso_date
            n_kpis += 1
            n_points += len(points)

    if latest_q:
        y, m = int(latest_q[:4]), int(latest_q[5:7])
        entry["quarter"] = f"{y}-Q{(m - 1) // 3 + 1}"
    else:
        entry["quarter"] = _current_quarter_label()
    print(f"ir_reports[{key}]: {n_kpis} KPIs aus Financials-XLS, {n_points} Datenpunkte gesamt "
          f"(Historie {earliest_q or '-'}..{latest_q or '-'})", flush=True)
    if n_kpis == 0:
        errors.append(("ir_reports", f"{name}: XLS gefunden, aber keine KPI-Zeilen erkannt (Struktur geaendert?) ({xlsx_url})"))


def _extract_pdf_kpis(pdf_bytes, patterns):
    import pdfplumber

    out = {}
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        text = "\n".join((p.extract_text() or "") for p in pdf.pages[:15])
    for kpi, pat in patterns.items():
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            try:
                out[kpi] = float(m.group(1).replace(",", "."))
            except ValueError:
                continue
    return out


def _fetch_pdf_company(data, key, comp, errors):
    name = comp.get("name", key)
    quarter = _current_quarter_label()
    reports = data.setdefault("ir_reports", {})
    pdf_url = _find_pdf(comp["ir_url"])
    entry = reports.setdefault(key, {})
    entry.update({
        "name": name, "quarter": quarter,
        "report_url": pdf_url or comp["ir_url"],
        "ir_url": comp["ir_url"], "checked": now_iso(),
        "scope": "konzern",
    })
    if not pdf_url:
        errors.append(("ir_reports", f"{name}: kein PDF-Link (JS-Seite?) - nur IR-Link gespeichert"))
        return
    patterns = comp.get("kpi_patterns") or {}
    if not patterns:
        return
    pdf_bytes = http_get(pdf_url, timeout=90).content
    kpis = _extract_pdf_kpis(pdf_bytes, patterns)
    entry["kpis"] = kpis
    qdate = quarter.replace("-Q1", "-03-31").replace("-Q2", "-06-30") \
                   .replace("-Q3", "-09-30").replace("-Q4", "-12-31")
    for kpi, val in kpis.items():
        sid = f"ir_{key}_" + "".join(c if c.isalnum() else "_" for c in kpi.lower())[:40]
        s = upsert_series(
            data, sid, label=f"{name}: {kpi}", frequency="quarterly",
            unit=kpi[kpi.find("(") + 1:kpi.find(")")] if "(" in kpi else "",
            scope="konzern",
            source=f"{name} Investor Relations (Quartalsbericht)",
            source_url=pdf_url,
        )
        add_point(s, qdate, val, "quarterly")
    if not kpis:
        errors.append(("ir_reports", f"{name}: PDF gefunden, aber keine KPIs per Regex extrahiert ({pdf_url})"))


def fetch(data, config, errors):
    for key, comp in config.get("ir_reports", {}).get("companies", {}).items():
        name = comp.get("name", key)
        try:
            if comp.get("financials_xlsx_title"):
                _fetch_xlsx_company(data, key, comp, errors)
            else:
                _fetch_pdf_company(data, key, comp, errors)
        except Exception as e:  # noqa: BLE001
            errors.append(("ir_reports", f"{name}: {e}"))
