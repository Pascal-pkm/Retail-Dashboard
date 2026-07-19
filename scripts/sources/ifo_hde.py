"""ifo-Geschaeftsklimaindex (absolute Punktwerte) + HDE-Konsumbarometer-Hinweise.

Primaerquelle (seit 07/2026): die von ifo direkt bereitgestellte Excel-Zeitreihe
unter https://www.ifo.de/ifo-zeitreihen ("ifo Geschaeftsklima (<Monat> <Jahr>)"),
die die komplette monatliche Historie seit 01/2005 enthaelt (Geschaeftsklima,
-lage, -erwartungen). Der Download-Link (Dateiname gsk-d-YYYYMM.xlsx) wechselt
jeden Monat, wir suchen ihn daher per Regex auf der Zeitreihen-Seite.

WICHTIG - Compliance-Hinweis (siehe COMPLIANCE.md Abschnitt 7): ifos eigene
"Bestellinformationen fuer ifo Zeitreihen" (verlinkt von der Download-Seite als
Nutzungsbedingung fuer genau diesen Excel-Service) erlauben die Nutzung der
Daten "nur zur privaten Information"; Weitergabe/Veroeffentlichung ist laut
ifo "nur nach besonderer Vereinbarung mit dem ifo Institut gestattet". Diese
Quelle wird hier trotzdem fuer die oeffentliche Website genutzt - eine bewusste,
vom Nutzer getroffene Entscheidung, das damit verbundene Risiko zu akzeptieren
(analog zur Hystreet-Abwaegung). Keine Weiterverbreitung der Originaldatei
selbst, nur einzelne extrahierte Zahlen mit Quellenangabe.

Fallback, falls der ifo-Download scheitert (Seitenstruktur geaendert, Layout-
Wechsel etc.): das bisherige Vorgehen ueber das Destatis "Dashboard Deutschland"
(Tile-Text der aktuellen Pressemitteilung, dl-de/by-2-0, rechtlich unkritisch,
aber nur 1-2 Punkte pro Lauf).
"""
import re
from datetime import date

from common import add_point, http_get, upsert_series
from sources.destatis_dashboard import get_tile, tile_widgets

TILE_IFO = "tile_1667288019608"
TILE_KONSUM = "tile_1667983271066"
SOURCE_IFO_DIRECT = "ifo Institut, ifo Zeitreihen (ifo-zeitreihen.de) - Nutzung ueber private-Info-Klausel hinaus akzeptiert, siehe COMPLIANCE.md"
SOURCE_DESTATIS = "ifo Institut via Destatis Dashboard Deutschland, dl-de/by-2-0"
ZEITREIHEN_URL = "https://www.ifo.de/ifo-zeitreihen"
GSK_XLSX_RE = re.compile(r'href="(/sites/default/files/secure/timeseries/gsk-d-\d{6}\.xlsx)"')

BROWSER_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
}

MONTHS = {"januar": 1, "februar": 2, "märz": 3, "april": 4, "mai": 5, "juni": 6,
          "juli": 7, "august": 8, "september": 9, "oktober": 10, "november": 11, "dezember": 12}

# Spalten im Sheet "ifo Geschäftsklima Deutschland" (Kopfzeile in Zeile 8, Daten ab Zeile 10)
GSK_COLUMNS = {
    2: ("ifo_geschaeftsklima", "ifo-Geschäftsklimaindex"),
    3: ("ifo_geschaeftslage", "ifo-Geschäftslage (Deutschland)"),
    4: ("ifo_geschaeftserwartungen", "ifo-Geschäftserwartungen (Deutschland)"),
}


def _strip_html(t):
    return re.sub(r"&nbsp;", " ", re.sub(r"<[^>]+>", " ", t or ""))


def _parse_version_date(s):
    """'März 2026' -> 2026-03-01."""
    m = re.match(r"(\w+)\s+(\d{4})", (s or "").strip())
    if m and m.group(1).lower() in MONTHS:
        return date(int(m.group(2)), MONTHS[m.group(1).lower()], 1).isoformat()
    return date.today().replace(day=1).isoformat()


def _find_gsk_xlsx_url():
    html = http_get(ZEITREIHEN_URL, headers=BROWSER_HEADERS).text
    m = GSK_XLSX_RE.search(html)
    if not m:
        return None
    return "https://www.ifo.de" + m.group(1)


def _fetch_ifo_direct(data, errors):
    """Primaerweg: komplette Historie seit 01/2005 direkt von ifo. Gibt True zurueck,
    wenn mindestens eine Serie erfolgreich befuellt wurde (dann kein Fallback noetig)."""
    import io
    import openpyxl

    xlsx_url = _find_gsk_xlsx_url()
    if not xlsx_url:
        errors.append(("ifo", "ifo-Geschäftsklima-XLSX-Link nicht gefunden (Seitenstruktur geändert?)"))
        return False

    xlsx_bytes = http_get(xlsx_url, headers=BROWSER_HEADERS, timeout=60).content
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    sheet_name = "ifo Geschäftsklima Deutschland"
    if sheet_name not in wb.sheetnames:
        errors.append(("ifo", f"Sheet '{sheet_name}' nicht in {xlsx_url} gefunden"))
        return False
    ws = wb[sheet_name]

    # Header-Zeile suchen (enthaelt "Monat/Jahr" in Spalte A) statt eine feste
    # Zeilennummer anzunehmen - robuster gegen kleinere Layout-Verschiebungen.
    header_row = None
    for r in range(1, min(15, ws.max_row) + 1):
        if str(ws.cell(r, 1).value or "").strip().lower() == "monat/jahr":
            header_row = r
            break
    if header_row is None:
        errors.append(("ifo", f"Header-Zeile 'Monat/Jahr' nicht gefunden in {xlsx_url}"))
        return False

    series_points = {sid: [] for sid, _ in GSK_COLUMNS.values()}
    for r in range(header_row + 1, ws.max_row + 1):
        raw_date = ws.cell(r, 1).value
        if not raw_date or not str(raw_date).strip():
            continue
        m = re.match(r"\s*(\d{2})/(\d{4})\s*$", str(raw_date))
        if not m:
            continue
        month, year = int(m.group(1)), int(m.group(2))
        when = date(year, month, 1).isoformat()
        for col, (sid, _label) in GSK_COLUMNS.items():
            val = ws.cell(r, col).value
            if isinstance(val, (int, float)):
                series_points[sid].append((when, float(val)))

    n_total = 0
    for col, (sid, label) in GSK_COLUMNS.items():
        pts = series_points[sid]
        if not pts:
            continue
        s = upsert_series(
            data, sid, label=label, frequency="monthly",
            unit="Punkte (2015=100, saisonbereinigt)", scope="branche",
            source=SOURCE_IFO_DIRECT, source_url=ZEITREIHEN_URL,
        )
        for when, val in pts:
            add_point(s, when, val, "monthly")
        n_total += len(pts)

    if n_total == 0:
        errors.append(("ifo", f"XLSX gefunden, aber keine Datenzeilen erkannt ({xlsx_url})"))
        return False

    print(f"ifo: {n_total} Datenpunkte aus direkter ifo-Zeitreihe geladen "
          f"({len(series_points['ifo_geschaeftsklima'])} Monate Geschäftsklima seit "
          f"{series_points['ifo_geschaeftsklima'][0][0] if series_points['ifo_geschaeftsklima'] else '?'})",
          flush=True)
    return True


def _fetch_ifo_via_destatis(data, errors):
    """Fallback: nur der/die im Destatis-Pressetext genannte(n) Punktwert(e)."""
    try:
        inner = get_tile(TILE_IFO)
        text = " ".join(_strip_html(c.get("text", "")) for c in inner.get("components", [])
                        if c.get("type") == "text")
        m = re.search(r"(?:stieg|sank|liegt|fiel|kletterte)\s+im\s+(\w+)\s+auf\s+(\d{2,3}[.,]\d)\s*Punkte", text)
        if not m:
            m = re.search(r"auf\s+(\d{2,3}[.,]\d)\s*Punkte", text)
            month_name = None
        else:
            month_name = m.group(1).lower()
        if not m:
            raise RuntimeError("kein Punktwert im Tile-Text gefunden (Struktur geaendert?)")
        value = float(m.group(m.lastindex).replace(",", "."))
        if not (50 <= value <= 130):
            raise RuntimeError(f"unplausibler Wert: {value}")

        when = _parse_version_date(inner.get("dataVersionDate", ""))
        year = int(when[:4])
        if month_name in MONTHS:
            when = date(year, MONTHS[month_name], 1).isoformat()

        s = upsert_series(
            data, "ifo_geschaeftsklima",
            label="ifo-Geschäftsklimaindex", frequency="monthly",
            unit="Punkte (2015=100)", scope="branche",
            source=SOURCE_DESTATIS, source_url="https://www.ifo.de/umfragen/ifo-geschaeftsklimaindex",
        )
        add_point(s, when, value, "monthly")

        m2 = re.search(r"nach\s+(\d{2,3}[.,]\d)\s*Punkten?\s+im\s+(\w+)", text)
        if m2 and month_name in MONTHS:
            prev_month_name = m2.group(2).lower()
            if prev_month_name in MONTHS:
                prev_value = float(m2.group(1).replace(",", "."))
                if 50 <= prev_value <= 130:
                    prev_year = year - 1 if MONTHS[prev_month_name] > MONTHS[month_name] else year
                    prev_when = date(prev_year, MONTHS[prev_month_name], 1).isoformat()
                    add_point(s, prev_when, prev_value, "monthly")

        return inner
    except Exception as e:  # noqa: BLE001
        errors.append(("ifo", str(e)))
        return None


def fetch(data, config, errors):
    widgets = []
    ok = False
    try:
        ok = _fetch_ifo_direct(data, errors)
    except Exception as e:  # noqa: BLE001
        errors.append(("ifo", f"Direkter ifo-Abruf fehlgeschlagen: {e}"))

    if not ok:
        inner = _fetch_ifo_via_destatis(data, errors)
        if inner:
            widgets.extend(tile_widgets(inner))

    # HDE-Konsumbarometer / GfK-Konsumklima: fertige Aussagen aus dem Konsum-Tile
    try:
        widgets.extend(tile_widgets(get_tile(TILE_KONSUM)))
    except Exception as e:  # noqa: BLE001
        errors.append(("hde_konsum", str(e)))

    data.setdefault("dashboard_widgets", {})["monthly"] = widgets
